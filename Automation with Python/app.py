import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#function to fix price to 10% less
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correct_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = correct_price

    values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(f"{filename}FIXED.xlsx")


#process_workbook("transactions.xlsx")